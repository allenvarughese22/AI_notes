import openpyxl
from openpyxl import Workbook
import openai
import os

# Set up OpenAI API key from environment variable
# openai.api_key = os.getenv("OPENAI_API_KEY")

openai_client = openai.OpenAI(
    api_key=os.getenv("OPENAI_API_KEY")  # Get API key from environment
)

# Check if API key is loaded
if not openai_client:
    raise ValueError("OPENAI_API_KEY environment variable not set.")

# Function to create or load an Excel workbook
def load_or_create_workbook(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        print(f"Loaded existing workbook: {filename}")
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(filename)
        print(f"Created a new workbook: {filename}")
    return workbook

# Function to generate Excel formula using GPT-4
def generate_formula(prompt):
    response = openai_client.chat.completions.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "You are an AI assistant that generates Excel formulas."},
            {"role": "user", "content": f"Generate an Excel formula for: {prompt}"}
        ],
        max_tokens=100,
        temperature=0.2
    )
    formula = response.choices[0].message.content.strip()
    return formula

# Function to apply the generated formula to a specified cell in Excel
def apply_formula(workbook, sheet_name, cell, formula):
    sheet = workbook[sheet_name]
    sheet[cell] = formula
    workbook.save(workbook.filename)
    print(f"Applied formula '{formula}' to cell {cell} in sheet '{sheet_name}'.")

# Main function to handle user input and workflow
def main():
    filename = "ai_generated_excel.xlsx"
    workbook = load_or_create_workbook(filename)
    sheet_name = "Sheet1"
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]

    print("Choose an option:")
    print("1. Generate Excel formula")
    print("2. Exit")
    choice = input("Enter 1 or 2: ")

    if choice == "1":
        prompt = input("Describe what formula you need (e.g., sum of A1 to A10): ")
        formula = generate_formula(prompt)
        print(f"Generated Formula: {formula}")
        cell = input("Enter the cell to apply this formula (e.g., B1): ")
        apply_formula(workbook, sheet_name, cell, formula)
    elif choice == "2":
        print("Exiting the application.")
    else:
        print("Invalid choice. Please select 1 or 2.")

if __name__ == "__main__":
    main()
